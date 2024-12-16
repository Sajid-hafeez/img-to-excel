import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
import io

def check_image_size(uploaded_file, max_size_mb=2):
    """Ensure the image file size is within the limit."""
    uploaded_file.seek(0, io.SEEK_END)
    file_size_mb = uploaded_file.tell() / (1024 * 1024)
    uploaded_file.seek(0)
    if file_size_mb > max_size_mb:
        raise ValueError(f"File size exceeds the maximum limit of {max_size_mb} MB.")

def image_to_excel(image):
    # Convert the resized image to RGB format
    img = image.convert("RGB")

    # Get image dimensions
    width, height = img.size

    # Create a new Excel workbook and active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Image Pixels"

    # Adjust cell width and height
    for col in range(1, width + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 4  # Adjust width for readability
    for row in range(1, height + 1):
        sheet.row_dimensions[row].height = 20  # Adjust height for readability

    # Streamlit progress number setup
    progress_text = st.empty()
    total_pixels = width * height

    # Iterate over the image pixels and set cell colors and values
    for y in range(height):
        for x in range(width):
            r, g, b = img.getpixel((x, y))
            hex_color = f"{r:02X}{g:02X}{b:02X}"  # Convert RGB to HEX format
            cell = sheet.cell(row=y + 1, column=x + 1)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell.value = f"{r},{g},{b}"  # Set cell value to RGB values

        # Update progress as a percentage
        if y % (height // 100 or 1) == 0:
            progress_text.text(f"Progress: {((y + 1) / height) * 100:.2f}%")

    # Save the workbook to a BytesIO stream
    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    progress_text.text("Progress: 100% Complete")
    return excel_stream

# Streamlit App with Modern Design
st.markdown(
    """<style>
    .main {
        background-color: #f5f5f5;
        font-family: Arial, sans-serif;
    }
    .stButton>button {
        background-color: #007BFF;
        color: white;
        border-radius: 5px;
        border: none;
        padding: 10px 20px;
        font-size: 16px;
        cursor: pointer;
        transition: background-color 0.3s;
    }
    .stButton>button:hover {
        background-color: #0056b3;
    }
    .stFileUploader {
        border: 2px dashed #007BFF;
        padding: 20px;
        background-color: #e9f5ff;
        border-radius: 10px;
    }
    .uploaded-image {
        border-radius: 10px;
        margin-top: 20px;
    }
    </style>""",
    unsafe_allow_html=True
)

st.title("🌟 Image to Excel Converter")
st.markdown("Convert your images into Excel sheets with pixel data and colors. Easy, fast, and elegant!")

uploaded_file = st.file_uploader(
    "Upload your image file (JPG, JPEG, PNG)",
    type=["jpg", "jpeg", "png"],
    label_visibility="visible"
)

if uploaded_file is not None:
    try:
        # Check if file size is within the limit
        check_image_size(uploaded_file, max_size_mb=2)

        # Load the image
        original_image = Image.open(uploaded_file)
        st.image(original_image, caption="Uploaded Image", use_container_width=True)

        # Generate Excel from the image
        with st.spinner("Processing the image... Please wait."):
            excel_data = image_to_excel(original_image)

        # Download link
        if st.download_button(
            label="📥 Download Excel File",
            data=excel_data,
            file_name="output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Click to download your generated Excel file."
        ):
            st.stop()

    except ValueError as e:
        st.error(str(e))
